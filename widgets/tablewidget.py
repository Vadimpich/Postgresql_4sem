from PySide6 import QtCore, QtWidgets
from PySide6.QtPrintSupport import QPrinter, QPrintDialog
from PySide6.QtWidgets import (
    QAbstractItemView, QTableWidgetItem, QFileDialog, QMessageBox
)
from openpyxl import Workbook

from dialogs import parentdialog
from modules.db import Database
from ui.widget_table import Ui_tableWidget


class TableWidget(QtWidgets.QWidget):
    def __init__(self, model, prev_index=None):
        super().__init__()
        self.ui = Ui_tableWidget()
        self.ui.setupUi(self)
        self.fields = []
        self.buttons = []
        self.model = model
        self.current_model = None
        self.db = Database()
        self.ui.tableName.setText(self.model.name)
        self.prev_index = prev_index
        self.setup_form()
        self.setup_table()
        self.update_table()
        self.setup_controls()

    def setup_form(self):
        for i in range(len(self.model.fields)):
            label = QtWidgets.QLabel()
            label.setText(self.model.fields[i].verbose)
            ref = self.model.fields[i].reference
            if ref:
                line = QtWidgets.QComboBox()
                items = self.db.select(
                    f'select {ref.table_name}_id, {ref.foreign_field} '
                    f'from {ref.table_name};'
                )
                for item in items:
                    line.addItem(f'{item[0]} | {item[1]}')
            else:
                line = QtWidgets.QLineEdit()
            line.setEnabled(self.model.fields[i].editable)
            self.fields.append(line)
            self.ui.formLayout.addRow(label, line)
        for i in range(len(self.model.buttons)):
            button = QtWidgets.QPushButton(self.model.buttons[i].verbose)
            button.clicked.connect(
                lambda x: self.exec_dialog(self.model.buttons[i].reference)
            )
            self.buttons.append(button)
            self.buttons[i].setDisabled(True)
            self.ui.formLayout.addRow(button)

    def exec_dialog(self, dialog_class):
        if self.current_model:
            dialog = dialog_class(self.current_model)
            res = dialog.exec()
            if res == QtWidgets.QDialog.DialogCode.Accepted:
                self.update_table()

    def setup_table(self):
        self.ui.dataTable.setColumnCount((len(self.model.fields) + 1))
        self.ui.dataTable.setHorizontalHeaderLabels(
            ['ID'] + [field.verbose for field in self.model.fields]
        )
        self.ui.dataTable.setColumnWidth(0, 50)
        for i in range(len(self.model.fields)):
            self.ui.dataTable.setColumnWidth(
                i + 1, 550 / len(self.model.fields) - 1
            )
        self.ui.dataTable.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection
        )
        self.ui.dataTable.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectItems
        )
        self.ui.dataTable.verticalHeader().setVisible(False)
        for field in self.model.fields:
            self.ui.searchBox.addItem(field.verbose)

    def setup_controls(self):
        self.ui.buttonAdd.clicked.connect(self.add)
        self.ui.buttonEdit.clicked.connect(self.edit)
        self.ui.buttonDelete.clicked.connect(self.delete)
        self.ui.dataTable.cellClicked.connect(self.select_row)
        self.ui.dataTable.cellDoubleClicked.connect(self.select_parent)
        self.ui.searchField.textChanged.connect(self.search)
        self.ui.saveButton.clicked.connect(self.save_table)
        self.ui.printButton.clicked.connect(self.print_table)
        if self.prev_index is not None:
            self.ui.buttonBack.clicked.connect(self.back)
        else:
            self.ui.buttonBack.hide()

    def update_table(self, search=None):
        data = []
        try:
            if not search:
                data = self.db.select(
                    f'select * from {self.model.view_name}'
                )
            else:
                field = (self.model.view_fields[
                    self.ui.searchBox.currentIndex() + 1]
                         if self.model.view_fields
                         else self.model.fields[
                    self.ui.searchBox.currentIndex()].name)
                data = self.db.select(
                    f'select * from {self.model.view_name} '
                    f'where CAST({field} AS VARCHAR) '
                    f"like '{search}%';"
                )
        except Exception as e:
            QMessageBox.warning(
                self, 'Ошибка', str(e), QMessageBox.StandardButton.Ok
            )
        self.ui.dataTable.setRowCount(len(data))
        for row in range(len(data)):
            for col in range(len(data[row])):
                item = QTableWidgetItem(f'{data[row][col]}')
                item.setFlags(
                    QtCore.Qt.ItemFlag.ItemIsSelectable
                    | QtCore.Qt.ItemFlag.ItemIsEnabled
                )
                self.ui.dataTable.setItem(row, col, item)

    def add(self):
        fields = [field.name for field in self.model.fields]
        values = [
            f"'{
                field.text()
                if isinstance(field, QtWidgets.QLineEdit)
                else field.currentText().split(' |')[0]
            }'"
            for field in self.fields
        ]
        try:
            self.db.execute(
                f'insert into {self.model.table_name}'
                f'({", ".join(fields)}) values ({", ".join(values)});'
            )
        except Exception as e:
            QMessageBox.warning(
                self, 'Ошибка', str(e), QMessageBox.StandardButton.Ok
            )
        self.update_table()

    def edit(self):
        row = self.ui.dataTable.currentRow()
        if row >= 0:
            if QtWidgets.QMessageBox.question(
                    self,
                    'Редактирование',
                    'Изменить ряд?',
                    QtWidgets.QMessageBox.StandardButton.Yes
                    | QtWidgets.QMessageBox.StandardButton.No
            ):
                try:
                    self.db.execute(
                        f'update {self.model.table_name} set '
                        f'{', '.join(
                            [f"{self.model.fields[i].name} = "
                             f"'{
                             self.fields[i].text()
                             if isinstance(self.fields[i], QtWidgets.QLineEdit)
                             else self.fields[i].currentText().split(' |')[0]
                             }'"
                             for i in range(len(self.model.fields))]
                        )}'
                        f'where {self.model.table_name}_id = '
                        f'{self.ui.dataTable.item(row, 0).text()};'
                    )
                except Exception as e:
                    QMessageBox.warning(
                        self, 'Ошибка', str(e), QMessageBox.StandardButton.Ok
                    )
            self.update_table()

    def delete(self):
        row = self.ui.dataTable.currentRow()
        if row >= 0:
            if QtWidgets.QMessageBox.question(
                    self,
                    'Удаление',
                    'Удалить ряд?',
                    QtWidgets.QMessageBox.StandardButton.Yes
                    | QtWidgets.QMessageBox.StandardButton.No
            ):
                try:
                    self.db.execute(
                        f'delete from {self.model.table_name} '
                        f'where {self.model.table_name}_id = '
                        f'{self.ui.dataTable.item(row, 0).text()};'
                    )
                except Exception as e:
                    QMessageBox.warning(
                        self, 'Ошибка', str(e), QMessageBox.StandardButton.Ok
                    )
            self.update_table()
            for field in self.fields:
                field.setText('')

    def select_row(self):
        row = self.ui.dataTable.currentRow()
        if row >= 0:
            self.current_model = [
                self.ui.dataTable.item(row, x).text()
                for x in range(self.ui.dataTable.columnCount())
            ]
            for i in range(len(self.fields)):
                if isinstance(self.fields[i], QtWidgets.QLineEdit):
                    self.fields[i].setText(
                        self.ui.dataTable.item(row, i + 1).text()
                    )
                else:
                    self.fields[i].setCurrentIndex(
                        self.find_index(
                            self.fields[i],
                            self.model.fields[i].reference,
                        )
                    )
            for button in self.buttons:
                button.setEnabled(True)
            self.ui.buttonEdit.setEnabled(True)
            self.ui.buttonDelete.setEnabled(True)
        else:
            self.current_model = None
            for button in self.buttons:
                button.setEnabled(False)
            self.ui.buttonEdit.setEnabled(False)
            self.ui.buttonDelete.setEnabled(False)

    def find_index(self, field, ref):
        row = self.ui.dataTable.currentRow()
        cell = self.ui.dataTable.item(row, 0).text()
        item_id = self.db.select(
            f'select {ref.table_name}_id '
            f'from {self.model.table_name} '
            f"where {self.model.table_name}_id = '{cell}'"
        )
        if item_id:
            item_id = item_id[0][0]
            for i in range(field.count()):
                if field.itemText(i).split(' |')[0] == str(item_id):
                    return i
        return -1

    def select_parent(self):
        ref = self.model.fields[
            self.ui.dataTable.currentColumn() - 1].reference
        if ref:
            parent_widget = TableWidget(
                ref
            )
            dialog = parentdialog.ParentDialog(parent_widget)
            dialog.exec()

    def search(self):
        if self.ui.searchBox.currentIndex() >= 0:
            self.update_table(self.ui.searchField.text())

    def print_table(self):
        printer = QPrinter()
        printer_dialog = QPrintDialog(printer)
        if printer_dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self.ui.dataTable.render(printer)

    def save_table(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл Excel", "",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        wb = Workbook()
        ws = wb.active
        rows = self.ui.dataTable.rowCount()
        cols = self.ui.dataTable.columnCount()
        for col in range(cols):
            header_item = self.ui.dataTable.horizontalHeaderItem(col)
            if header_item is not None:
                ws.cell(row=1, column=col + 1, value=str(header_item.text()))
        for row in range(rows):
            for col in range(cols):
                item = self.ui.dataTable.item(row, col)
                if item is not None:
                    ws.cell(row=row + 2, column=col + 1,
                            value=str(item.text()))
        wb.save(file_path)

    def back(self):
        self.parent().setCurrentIndex(self.prev_index)
        self.parent().removeWidget(self)
        del self
